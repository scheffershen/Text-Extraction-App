import os
import zipfile
from docx import Document
from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import load_workbook
from extract_msg import Message
from PIL import Image
import pytesseract
from io import BytesIO
from xlrd import open_workbook
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# Function to extract text from different document types


def extract_text_from_file(file_path):
    file_ext = os.path.splitext(file_path)[1].lower()
    extracted_text = ""

    if file_ext == ".docx":
        doc = Document(file_path)
        extracted_text = "\n".join([para.text for para in doc.paragraphs])

    elif file_ext == ".pdf":
        reader = PdfReader(file_path)
        extracted_text = "\n".join([page.extract_text()
                                   for page in reader.pages])

    elif file_ext in [".xlsx", ".xlsm"]:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        extracted_text = "\n".join([str(cell.value)
                                   for row in sheet.iter_rows() for cell in row])

    elif file_ext == ".xls":
        workbook = open_workbook(file_path)
        for sheet in workbook.sheets():
            for row in range(sheet.nrows):
                extracted_text += " ".join([str(sheet.cell_value(row, col))
                                           for col in range(sheet.ncols)]) + "\n"

    elif file_ext == ".msg":
        msg = Message(file_path)
        extracted_text = f"Subject: {msg.subject}\nFrom: {msg.sender}\nTo: {msg.to}\nCC: {msg.cc}\nDate: {msg.date}\n\n{msg.body}"

    elif file_ext == ".txt":
        with open(file_path, 'r', encoding='utf-8') as file:
            extracted_text = file.read()

    elif file_ext in [".png", ".jpg", ".jpeg"]:
        image = Image.open(file_path)
        extracted_text = pytesseract.image_to_string(image)

    elif file_ext == ".zip":
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            extracted_text = ""
            with zip_ref.open(zip_ref.namelist()[0]) as file:
                extracted_text = extract_text_from_file(BytesIO(file.read()))

    elif file_ext == ".doc":
        try:
            import subprocess
            extracted_text = subprocess.check_output(
                ['antiword', file_path]).decode('utf-8')
        except FileNotFoundError:
            print(
                "antiword not installed. Install it or use another method to process .doc files.")

    return extracted_text

# Function to generate a unique file name if a file with the same name already exists


def get_unique_file_path(directory, base_filename, extension):
    counter = 1
    output_file_path = os.path.join(directory, f"{base_filename}{extension}")
    while os.path.exists(output_file_path):
        output_file_path = os.path.join(
            directory, f"{base_filename}({counter}){extension}")
        counter += 1
    return output_file_path


# Scan directory and process files
directory = "./input"
output_directory = "./output"

if not os.path.exists(output_directory):
    os.makedirs(output_directory)

files_processed = False

for filename in os.listdir(directory):
    file_path = os.path.join(directory, filename)
    if os.path.isfile(file_path):
        files_processed = True
        extracted_text = extract_text_from_file(file_path)

        # Generate a unique file path to avoid overwriting
        base_filename = os.path.splitext(filename)[0]
        output_file_path = get_unique_file_path(
            output_directory, base_filename, ".txt")

        # Save the extracted text to a txt file in the output directory
        with open(output_file_path, 'w', encoding='utf-8') as output_file:
            output_file.write(extracted_text)

if not files_processed:
    print(f"No files found in the directory: {directory}")
